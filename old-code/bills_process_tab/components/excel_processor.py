# src/components/tabs/bills_process_tab/components/excel_processor.py
import os
from typing import List, Any, Optional
from openpyxl import load_workbook
from src.components.common.custom_messagebox import CustomMessageBox
from .result import Result
from .excel_components.state_manager import StateManager
from .excel_components.date_formatter import DateFormatter
from .excel_components.worksheet_manager import WorksheetManager
from .excel_components.formula_manager import FormulaManager
from .excel_components.holiday_validator import HolidayValidator

class ExcelProcessor:
    """
    Handles Excel template processing and data insertion.
    Main orchestrator class that coordinates various Excel processing components.
    """
    
    def __init__(self):
        self.state_manager = StateManager()
        self.date_formatter = DateFormatter()
        self.worksheet_manager = WorksheetManager()
        self.formula_manager = FormulaManager()
        self.holiday_validator = HolidayValidator()
    
    def process_template(self, template_file_path: Optional[str], day: int, daily_files: List[str], 
                        collected_data: List[List[Any]], inventory_code: str, target_dir: str, 
                        parent=None, year: int = None, month: int = None) -> Result:
        """
        Process the Excel template by inserting data from daily files.
        
        Args:
            template_file_path: Path to the template Excel file
            day: Day number (used as sheet name)
            daily_files: List of daily file names
            collected_data: Matrix of collected data from daily files
            inventory_code: Inventory code from configuration
            target_dir: The directory containing the daily files.
            parent: Parent widget for error dialogs
            year: Year for date formatting (optional)
            month: Month for date formatting (optional)
            
        Returns:
            Result indicating success or failure
        """
        if not template_file_path or not os.path.exists(template_file_path):
            return Result.error("❌ ไม่พบไฟล์เทมเพลตสำหรับประมวลผล Excel.")

        # Load processing state
        state_file_path = os.path.join(target_dir, ".processing_state.json")
        processed_files = self.state_manager.load_processed_files(state_file_path)

        # Early exit if all files are already processed
        if self._check_early_exit(daily_files, processed_files, parent):
            return Result.success("✅ ไฟล์ทั้งหมดได้รับการประมวลผลแล้ว.")

        # Filter out files that have already been processed
        new_daily_files = [f for f in daily_files if f not in processed_files]
        
        if not new_daily_files:
            return Result.success("ไม่มีไฟล์ใหม่ให้ประมวลผล.")

        # Get corresponding collected data for new files
        new_collected_data = self._get_corresponding_data(daily_files, new_daily_files, collected_data)
        
        try:
            # Load workbook and get worksheet
            wb = load_workbook(template_file_path)
            sheet_name = str(day)
            
            if sheet_name not in wb.sheetnames:
                return Result.error(f"❌ ไม่พบชีทชื่อ '{sheet_name}' ในไฟล์เทมเพลต.")
            
            ws = wb[sheet_name]

            # Calculate processing parameters
            num_existing_files = len(processed_files)
            num_new_files = len(new_daily_files)
            first_new_row = 5 + num_existing_files
            total_files = num_existing_files + num_new_files

            # Fill processing date
            if year and month:
                thai_date = self.date_formatter.format_thai_date(day, month, year)
                ws['D2'] = thai_date
                ws['D2'].font = self.date_formatter.date_font
            
            # Prepare worksheet structure
            self.worksheet_manager.prepare_worksheet(ws, num_new_files, num_existing_files)
            
            # Fill data for new files
            self.worksheet_manager.fill_data(ws, new_daily_files, new_collected_data, inventory_code, start_row=first_new_row)

            # Add formulas
            first_data_row = 5
            last_data_row = first_data_row + total_files - 1
            self.formula_manager.add_sum_formulas(ws, first_data_row, last_data_row)
            
            # Handle cumulative sales formula
            self._handle_cumulative_sales(wb, ws, day, last_data_row)

            # Save workbook
            wb.save(template_file_path)

            # Update state
            self.state_manager.save_processed_files(
                state_file_path, 
                processed_files.union(set(new_daily_files))
            )
            
            return Result.success("อัปเดตไฟล์เทมเพลตเรียบร้อยแล้ว.")
            
        except PermissionError:
            if parent:
                CustomMessageBox.show_error(
                    parent,
                    "ข้อผิดพลาดการเข้าถึงไฟล์",
                    "ไฟล์เทมเพลตกำลังเปิดใช้งานอยู่ กรุณาปิดไฟล์ก่อนประมวลผล."
                )
            return Result.error("❌ ไฟล์เทมเพลตกำลังเปิดใช้งานอยู่ กรุณาปิดไฟล์ก่อนประมวลผล.")
        
        except Exception as e:
            if parent:
                CustomMessageBox.show_error(
                    parent,
                    "ข้อผิดพลาดการประมวลผล",
                    f"เกิดข้อผิดพลาดขณะประมวลผลไฟล์เทมเพลต: {e}"
                )
            return Result.error(f"❌ เกิดข้อผิดพลาดขณะประมวลผลไฟล์เทมเพลต: {e}")

    def _check_early_exit(self, daily_files: List[str], processed_files: set, parent) -> bool:
        """Check if we should exit early (all files already processed)."""
        if all(f in processed_files for f in daily_files):
            if processed_files:  # Only show if there are actually processed files
                if parent:
                    CustomMessageBox.show_info(
                        parent,
                        "ประมวลผลแล้ว",
                        "ไฟล์ทั้งหมดได้รับการประมวลผลแล้วสำหรับวันนี้."
                    )
                return True
        return False

    def _get_corresponding_data(self, daily_files: List[str], new_daily_files: List[str], 
                              collected_data: List[List[Any]]) -> List[List[Any]]:
        """Get collected data corresponding to new daily files."""
        new_collected_data = []
        for i, f in enumerate(daily_files):
            if f in new_daily_files:
                new_collected_data.append(collected_data[i])
        return new_collected_data

    def _handle_cumulative_sales(self, workbook, worksheet, day: int, last_data_row: int):
        """Handle the cumulative sales formula logic."""
        # Find cumulative sales row
        row_yodkaysum = None
        value_yodruam = None
        
        for r in range(1, worksheet.max_row + 1):
            d_val = worksheet[f"D{r}"].value
            if d_val == "ยอดขายสะสม":
                row_yodkaysum = r
            if d_val == "ยอดรวม":
                value_yodruam = worksheet[f"H{r}"].value

        if not row_yodkaysum:
            return

        if day == 1:
            # First day: just use current day's total
            if value_yodruam is not None:
                worksheet[f"H{row_yodkaysum}"] = value_yodruam
        else:
            # Find last valid working day
            valid_prev_day, valid_prev_row = self.holiday_validator.find_last_valid_working_day(workbook, day)
            
            if valid_prev_day and valid_prev_row:
                # Found valid previous working day
                worksheet[f"H{row_yodkaysum}"] = f"=SUM(H5:H{last_data_row})+'{valid_prev_day}'!H{valid_prev_row}"
            else:
                # No valid previous day found, just use current sum
                worksheet[f"H{row_yodkaysum}"] = f"=SUM(H5:H{last_data_row})"