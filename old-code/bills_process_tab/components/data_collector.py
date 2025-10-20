# src/components/tabs/bills_process_tab/components/data_collector.py
import os
from typing import List, Any, Tuple
import re
from openpyxl import load_workbook
from src.components.settings.bills.models import BillsConfigManager
from .result import Result

class DataCollector:
    """
    Handles collecting data from daily bill files based on configuration rules.
    """
    
    def __init__(self):
        self.config_mgr = BillsConfigManager()
    
    def collect_from_files(self, daily_files: List[str], target_dir: str) -> Result:
        """
        Collect data from all daily files based on bill configuration.
        
        Args:
            daily_files: List of daily bill file names
            target_dir: Directory containing the daily files
            
        Returns:
            Result: Contains collected data matrix on success
        """
        try:
            # Load bill configuration
            bill_items = self.config_mgr.load_config()
            
            collected_data = []
            for fname in daily_files:
                file_path = os.path.join(target_dir, fname)
                file_result = self._collect_from_single_file(file_path, bill_items)
                collected_data.append(file_result)
            
            return Result.success(collected_data)
            
        except Exception as e:
            return Result.error(f"❌ เกิดข้อผิดพลาดขณะดึงข้อมูลจากไฟล์: {e}")
    
    def _collect_from_single_file(self, file_path: str, bill_items: List[Any]) -> List[Any]:
        """
        Collect data from a single Excel file based on bill item configurations.
        
        Args:
            file_path: Path to the Excel file
            bill_items: List of bill configuration items
            
        Returns:
            List of collected values with phone numbers processed (empty strings on error)
        """
        try:
            wb_daily = load_workbook(file_path, data_only=True)
            ws_daily = wb_daily.active
            
            result = []
            phone_number = ""
            
            for idx, item in enumerate(bill_items):
                value = self._extract_value_by_type(ws_daily, item)
                
                # Process the value if it contains address/phone data (typically for column E)
                if isinstance(value, str) and value.strip():
                    processed_value, extracted_phone = self._process_address_phone_data(value) if idx == 1 else (value, "")
                    result.append(processed_value)
                    
                    # Store phone number for later use (will be added to result for column M)
                    if extracted_phone:
                        phone_number = extracted_phone
                else:
                    result.append(value)
            
            # Add the extracted phone number as an additional item for column M (6th value)
            result.append(phone_number)
            
            # Detect Lazada/Shopee in config id 5 value (index 4)
            platform = ""
            if len(result) > 4 and isinstance(result[4], str):
                val_lower = result[4].lower()
                if "lazada" in val_lower:
                    platform = "Lazada"
                elif "shopee" in val_lower:
                    platform = "Shopee"
            result.append(platform)  # For column L
            
            return result
            
        except Exception as e:
            # Return empty values for this file if there's an error
            # Include extra empty string for phone column
            return [""] * (len(bill_items) + 1)
    
    def _process_address_phone_data(self, data: str) -> Tuple[str, str]:
        """
        Process address/phone data by extracting phone number and formatting address.
        
        Args:
            data: Raw data containing address and phone information
            
        Returns:
            Tuple of (formatted_address, formatted_phone_number)
        """
        # Remove "Tel. ", "โทร.", "Tel. โทร." prefixes and extract phone number
        phone_pattern = r'(?:Tel\.\s*(?:โทร\.)?\s*|โทร\.\s*)([0-9\-]{9,12})'
        phone_match = re.search(phone_pattern, data)
        
        extracted_phone = ""
        cleaned_data = data
        
        if phone_match:
            # Extract and format phone number
            raw_phone = phone_match.group(1)
            extracted_phone = self._format_phone_number(raw_phone)
            
            # Remove the phone part from the original data
            cleaned_data = re.sub(r'(?:Tel\.\s*(?:โทร\.)?\s*|โทร\.\s*)[0-9\-]+', '', data)
        
        # Extract province and zone from address
        formatted_address = self._extract_province_zone(cleaned_data)
        
        return formatted_address, extracted_phone
    
    def _extract_province_zone(self, address: str) -> str:
        """
        Extract province and zone from Thai address format.
        
        Args:
            address: Address string containing จ.{province} and ต.{zone}
            
        Returns:
            Formatted string as "province / zone"
        """
        # Clean up the address data - remove extra whitespace
        cleaned_address = re.sub(r'\s+', ' ', address.strip())
        
        # Extract province (จ.{province})
        province_match = re.search(r'จ\.([^\s]+)', cleaned_address)
        province = province_match.group(1) if province_match else ""
        
        # Extract zone/district (ต.{zone})  
        zone_match = re.search(r'ต\.([^\s]+)', cleaned_address)
        zone = zone_match.group(1) if zone_match else ""
        
        # Format as "province / zone"
        if province and zone:
            return f"{province} / {zone}"
        elif province:
            return province
        elif zone:
            return zone
        else:
            # If no จ. or ต. found, return cleaned address as fallback
            return cleaned_address
    
    def _format_phone_number(self, phone: str) -> str:
        """
        Format phone number to xxx-xxx-xxxx (10 digits) or xx-xxx-xxxx (9 digits).
        
        Args:
            phone: Raw phone number (may contain dashes or be plain digits)
            
        Returns:
            Formatted phone number
        """
        # Remove all non-digit characters
        digits_only = re.sub(r'[^0-9]', '', phone)
        
        if len(digits_only) == 10:
            # Format as xxx-xxx-xxxx
            return f"{digits_only[:3]}-{digits_only[3:6]}-{digits_only[6:]}"
        elif len(digits_only) == 9:
            # Format as xx-xxx-xxxx
            return f"{digits_only[:2]}-{digits_only[2:5]}-{digits_only[5:]}"
        else:
            # Return as-is if length doesn't match expected formats
            return phone
    
    def _extract_value_by_type(self, worksheet, item) -> Any:
        """
        Extract value from worksheet based on item type and configuration.
        
        Args:
            worksheet: Openpyxl worksheet object
            item: Bill configuration item
            
        Returns:
            Extracted value or empty string
        """
        try:
            if item.type == 0:
                # By Cell: check cell value, then collect from focus cell
                return self._extract_by_cell(worksheet, item)
            elif item.type == 2:
                # By Collect Col: scan focus column for check value, collect from value column
                return self._extract_by_collect_col(worksheet, item)
            else:
                return ""
        except Exception:
            return ""
    
    def _extract_by_cell(self, worksheet, item) -> Any:
        """Extract value using 'By Cell' method."""
        check_val = worksheet[item.check].value if item.check else None
        if check_val == item.value:
            focus_val = worksheet[item.focus].value if item.focus else ""
            return focus_val
        return ""
    
    def _extract_by_collect_col(self, worksheet, item) -> Any:
        """Extract value using 'By Collect Col' method."""
        focus_col = item.focus
        check_val = item.check
        collect_col = item.value
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            cell_focus = None
            cell_collect = None
            
            for cell in row:
                if cell.column_letter == focus_col:
                    cell_focus = cell
                if cell.column_letter == collect_col:
                    cell_collect = cell
            
            if cell_focus and cell_focus.value == check_val:
                return cell_collect.value if cell_collect else ""
        
        return ""