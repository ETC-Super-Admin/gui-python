import os
from openpyxl import load_workbook
from src.components.common.custom_messagebox import CustomMessageBox, MessageType

def update_formula(year, month, day, parent=None):
    """
    Scan column E for cells containing 'Lazada' or 'Shopee' and print their positions to the left panel.
    """
    from src.components.settings.bills.models import BillsConfigManager
    config_mgr = BillsConfigManager()
    app_dir = config_mgr.app_dir
    file1_path = os.path.join(app_dir, "template_path_settings.json")
    import json
    try:
        with open(file1_path, "r", encoding="utf-8") as f:
            config = json.load(f)
            path1 = config.get("path1", None)
    except Exception:
        path1 = None
    if not path1:
        msg = "❌ ไม่พบเส้นทางไฟล์แม่แบบ"
        style = CustomMessageBox.get_text_display_style(MessageType.ERROR)
        return msg, "error"
    folder = os.path.join(path1, f"Year_{year:04d}", f"Month_{month:02d}")
    # Find the first .xlsx file in the folder
    template_file = None
    if os.path.exists(folder):
        for fname in os.listdir(folder):
            if fname.lower().endswith('.xlsx'):
                template_file = os.path.join(folder, fname)
                break
    if not template_file or not os.path.exists(template_file):
        msg = "❌ ไม่พบไฟล์แม่แบบ (.xlsx) ในโฟลเดอร์ที่กำหนด"
        style = CustomMessageBox.get_text_display_style(MessageType.ERROR)
        return msg, "error"
    try:
        wb = load_workbook(template_file)
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            msg = f"❌ ไม่พบชีท '{sheet_name}' ในไฟล์แม่แบบ"
            style = CustomMessageBox.get_text_display_style(MessageType.ERROR)
            return msg, "error"
        ws = wb[sheet_name]
        lazada_rows = []
        shopee_rows = []
        # Count number of bill rows (non-empty in column E, starting from row 5)
        bill_rows = [row[0] for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5) if row[0].value]
        num_bills = len(bill_rows)
        if num_bills > 0:
            last_data_row = 5 + num_bills - 1
        else:
            last_data_row = 6  # fallback if no data
        # Read processed files from .processing_state.json to determine last_data_row
        cache_file = os.path.join(folder, "Daily_Bills", f"Day_{day}", ".processing_state.json")
        num_processed_files = 0
        if os.path.exists(cache_file):
            try:
                with open(cache_file, "r", encoding="utf-8") as f:
                    state = json.load(f)
                    if isinstance(state, list):
                        num_processed_files = len(state)
                    else:
                        num_processed_files = len(state.get("processed_files", []))
            except Exception:
                pass
        if num_processed_files > 0:
            last_data_row = 5 + num_processed_files - 1
        else:
            last_data_row = 6  # fallback if no data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
            cell = row[0]
            val = str(cell.value) if cell.value is not None else ""
            if "lazada" in val.lower():
                lazada_rows.append((cell.row, val))
                ws[f"F{cell.row}"] = f"=COUNTIF(L5:L{last_data_row},\"Lazada\")"
                ws[f"G{cell.row}"] = f"=SUMIF(L5:L{last_data_row},\"Lazada\",G5:G{last_data_row})"
                ws[f"H{cell.row}"] = f"=SUMIF(L5:L{last_data_row},\"Lazada\",H5:H{last_data_row})"
            if "shopee" in val.lower():
                shopee_rows.append((cell.row, val))
                ws[f"F{cell.row}"] = f"=COUNTIF(L5:L{last_data_row},\"Shopee\")"
                ws[f"G{cell.row}"] = f"=SUMIF(L5:L{last_data_row},\"Shopee\",G5:G{last_data_row})"
                ws[f"H{cell.row}"] = f"=SUMIF(L5:L{last_data_row},\"Shopee\",H5:H{last_data_row})"
        from src.components.tabs.bills_process_tab.components.excel_components.holiday_validator import HolidayValidator
        holiday_validator = HolidayValidator()
        # After filling Lazada/Shopee formulas, find 'Grand total' in column E
        row_grand_total = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
            cell = row[0]
            val = str(cell.value) if cell.value is not None else ""
            if val.strip().lower() == "grand total":
                row_grand_total = cell.row
                break
        # Find Shopee and Lazada row numbers for formula range
        def is_valid_cumulative_value(cell_value):
            if cell_value in (None, "", 0, "=0"):
                return False
            if isinstance(cell_value, str):
                try:
                    num_val = float(cell_value)
                    return num_val > 0
                except (ValueError, TypeError):
                    return bool(cell_value.strip())
            if isinstance(cell_value, (int, float)):
                return cell_value > 0
            return False
        def find_last_valid_e_row(workbook, current_day, keyword):
            prev_day = current_day - 1
            while prev_day >= 1:
                prev_sheet_name = str(prev_day)
                if prev_sheet_name not in workbook.sheetnames:
                    prev_day -= 1
                    continue
                ws_prev = workbook[prev_sheet_name]
                last_row = None
                for row in ws_prev.iter_rows(min_row=1, max_row=ws_prev.max_row, min_col=5, max_col=5):
                    cell = row[0]
                    val = str(cell.value) if cell.value is not None else ""
                    if keyword.lower() in val.lower():
                        # Check if column I of this row is valid
                        i_val = ws_prev[f"I{cell.row}"].value
                        if is_valid_cumulative_value(i_val):
                            last_row = cell.row
                if last_row:
                    return prev_sheet_name, last_row
                prev_day -= 1
            return None, None
        if shopee_rows and lazada_rows and row_grand_total:
            row_found_shopee = shopee_rows[0][0]
            row_found_lazada = lazada_rows[-1][0]
            ws[f"F{row_grand_total}"] = f"=SUM(F{row_found_shopee}:F{row_found_lazada})"
            ws[f"G{row_grand_total}"] = f"=SUM(G{row_found_shopee}:G{row_found_lazada})"
            ws[f"H{row_grand_total}"] = f"=SUM(H{row_found_shopee}:H{row_found_lazada})"
            if str(day) == "1":
                ws[f"I{row_found_shopee}"] = f"=H{row_found_shopee}"
                ws[f"I{row_found_lazada}"] = f"=H{row_found_lazada}"
            else:
                prev_shopee_day, prev_shopee_row = find_last_valid_e_row(wb, int(day), "Shopee")
                prev_lazada_day, prev_lazada_row = find_last_valid_e_row(wb, int(day), "Lazada")
                if prev_shopee_day and prev_shopee_row:
                    ws[f"I{row_found_shopee}"] = f"=H{row_found_shopee}+'{prev_shopee_day}'!I{prev_shopee_row}"
                else:
                    ws[f"I{row_found_shopee}"] = f"=H{row_found_shopee}"
                if prev_lazada_day and prev_lazada_row:
                    ws[f"I{row_found_lazada}"] = f"=H{row_found_lazada}+'{prev_lazada_day}'!I{prev_lazada_row}"
                else:
                    ws[f"I{row_found_lazada}"] = f"=H{row_found_lazada}"
            ws[f"I{row_grand_total}"] = f"=SUM(I{row_found_shopee}:I{row_found_lazada})"
        wb.save(template_file)
        msg_lines = [f"🔎 พบ Lazada ในคอลัมน์ E ที่แถว:"]
        if lazada_rows:
            for r, v in lazada_rows:
                msg_lines.append(f"  - แถว {r}: {v}")
        else:
            msg_lines.append("  (None)")
        msg_lines.append("")
        msg_lines.append(f"🔎 พบ Shopee ในคอลัมน์ E ที่แถว:")
        if shopee_rows:
            for r, v in shopee_rows:
                msg_lines.append(f"  - แถว {r}: {v}")
        else:
            msg_lines.append("  (None)")
        msg = "\n".join(msg_lines)
        style = CustomMessageBox.get_text_display_style(MessageType.INFO)
        return msg, "info"
    except PermissionError:
        msg = "❌ ไฟล์แม่แบบกำลังถูกเปิดใช้งานอยู่ กรุณาปิดไฟล์ก่อนทำการประมวลผล"
        style = CustomMessageBox.get_text_display_style(MessageType.ERROR)
        if parent:
            CustomMessageBox.show_error(parent, "การเข้าถึงไฟล์ล้มเหลว", msg)
        return msg, "error"
    except Exception as e:
        msg = f"❌ เกิดข้อผิดพลาดในการค้นหา Lazada/Shopee: {e}"
        style = CustomMessageBox.get_text_display_style(MessageType.ERROR)
        return msg, "error"
