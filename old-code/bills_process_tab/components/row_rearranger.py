# src/components/tabs/bills_process_tab/components/row_rearranger.py
import os
from openpyxl import load_workbook
from src.components.common.custom_messagebox import CustomMessageBox

def rearrange_template_rows(year, month, day, file1_path, parent=None):
    """
    Rearrange rows in the template Excel file for the given date based on column E (province/zone).
    Args:
        year, month, day: Date components
        file1_path: Path to template_path_settings.json
        parent: Parent widget for message boxes
    Returns:
        (message, message_type)
    """
    import json
    try:
        # Load config to get path1
        if not os.path.exists(file1_path):
            return ("❌ ไม่พบไฟล์กำหนดค่า Path1.", "error")
        with open(file1_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        path1 = config.get("path1", None)
        if not path1:
            return ("❌ Path1 ยังไม่ได้ถูกกำหนดค่า.", "error")
        # Locate template file
        month_dir = os.path.join(path1, f"Year_{year:04d}", f"Month_{month:02d}")
        if not os.path.exists(month_dir):
            return (f"❌ ไม่พบโฟลเดอร์ของเดือน: {month_dir}", "error")
        template_file = None
        for fname in os.listdir(month_dir):
            if fname.lower().endswith('.xlsx'):
                template_file = os.path.join(month_dir, fname)
                break
        if not template_file or not os.path.exists(template_file):
            return ("❌ ไม่พบไฟล์เทมเพลต.", "error")
        # Open workbook and sheet
        wb = load_workbook(template_file)
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            return (f"❌ ไม่พบชีทชื่อ '{sheet_name}' ในไฟล์เทมเพลต.", "error")
        ws = wb[sheet_name]
        # Read all data rows (starting from row 5, until first empty A cell or until formulas row)
        data_rows = []
        row = 5
        while True:
            a_val = ws[f"A{row}"].value
            e_val = ws[f"E{row}"].value
            # Stop if A is empty or if we hit a formula row (A cell is a formula)
            if a_val is None or (isinstance(a_val, str) and a_val.startswith("=")):
                break
            # Read all columns A-N (1-14)
            row_data = [ws.cell(row=row, column=col).value for col in range(1, 15)]
            data_rows.append(row_data)
            row += 1
        if not data_rows:
            return ("ไม่มีข้อมูลแถวที่จะจัดเรียงใหม่.", "info")
        # Sort by column L (index 11), then by column E (index 4)
        def l_sort_key(row):
            l_val = (row[11] or '').strip().lower() if row[11] else ''
            if l_val == 'lazada':
                l_order = 0
            elif l_val == 'shopee':
                l_order = 1
            else:
                l_order = 2
            e_val = str(row[4]) if row[4] is not None else ''
            return (l_order, e_val)
        data_rows.sort(key=l_sort_key)
        # Write back sorted rows and update index in column A
        for idx, row_data in enumerate(data_rows):
            row_num = 5 + idx
            row_data[0] = idx + 1  # Update column A (index) to be sequential
            for col in range(1, 15):
                ws.cell(row=row_num, column=col).value = row_data[col-1]
        wb.save(template_file)
        if parent:
            CustomMessageBox.show_success(parent, "จัดเรียงข้อมูลสำเร็จ", "ได้ทำการจัดเรียงข้อมูลตามจังหวัด/โซนแล้ว.")
        return ("✅ จัดเรียงข้อมูลตามจังหวัด/โซนเรียบร้อยแล้ว.", "success")
    except Exception as e:
        if parent:
            CustomMessageBox.show_error(parent, "เกิดข้อผิดพลาดในการจัดเรียง", str(e))
        return (f"❌ เกิดข้อผิดพลาดในการจัดเรียงข้อมูล: {e}", "error")
