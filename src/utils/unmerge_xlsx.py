import os
from openpyxl import load_workbook

def unmerge_cells_in_file(file_path: str) -> bool:
    """
    Unmerges all merged cells in all worksheets of an Excel file and saves it.
    
    Args:
        file_path: The path to the Excel file.
        
    Returns:
        True if successful, False otherwise.
    """
    if not os.path.exists(file_path):
        return False

    try:
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            if sheet.merged_cells:
                # Create a list of merged cell ranges to iterate over
                merged_ranges = list(sheet.merged_cells)
                for merged_cell_range in merged_ranges:
                    sheet.unmerge_cells(str(merged_cell_range))
        wb.save(file_path)
        return True
    except PermissionError:
        print(f"Permission denied: The file is open or you don't have write access: {file_path}")
        return False
    except Exception as e:
        print(f"Error unmerging cells in {file_path}: {e}")
        return False
