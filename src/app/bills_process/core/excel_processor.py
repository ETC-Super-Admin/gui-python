import os
from openpyxl import load_workbook
from .result import Result
from .excel_components.state_manager import StateManager
from .excel_components.worksheet_manager import WorksheetManager
from .excel_components.date_formatter import DateFormatter
from .excel_components.formula_manager import FormulaManager
from .excel_components.holiday_validator import HolidayValidator

class ExcelProcessor:
    def __init__(self, template_path, data, date, inventory_code, target_dir, daily_files):
        self.template_path = template_path
        self.data = data
        self.date = date
        self.inventory_code = inventory_code
        self.target_dir = target_dir
        self.daily_files = daily_files
        self.state_manager = StateManager()
        self.worksheet_manager = WorksheetManager()
        self.date_formatter = DateFormatter()
        self.formula_manager = FormulaManager()
        self.holiday_validator = HolidayValidator()

    def process(self, **kwargs) -> Result:
        """
        Main method to process the Excel template.
        """
        progress_callback = kwargs.get('progress_callback')
        def report_progress(msg):
            if progress_callback:
                progress_callback.emit(msg)

        state_file_path = os.path.join(self.target_dir, ".processing_state.json")
        
        try:
            # 1. Load processing state
            report_progress("  > กำลังโหลดสถานะการประมวลผล...")
            processed_files_set = self.state_manager.load_processed_files(state_file_path)
            report_progress(f"  > พบไฟล์ที่ประมวลผลแล้ว {len(processed_files_set)} ไฟล์")

            # 2. Filter for new files
            new_daily_files = [f for f in self.daily_files if f not in processed_files_set]

            if not new_daily_files:
                return Result.info("✅ ไฟล์บิลรายวันทั้งหมดสำหรับวันนี้ได้รับการประมวลผลแล้ว")
            
            report_progress(f"  > พบไฟล์ใหม่ {len(new_daily_files)} ไฟล์ที่ต้องประมวลผล")

            new_data = []
            for i, f in enumerate(self.daily_files):
                if f in new_daily_files:
                    new_data.append(self.data[i])

            # 3. Load workbook
            report_progress("  > กำลังโหลดสมุดงาน...")
            wb = load_workbook(self.template_path)
            sheet_name = str(self.date.day())
            
            if sheet_name not in wb.sheetnames:
                return Result.error(f"❌ ไม่พบชีทชื่อ '{sheet_name}' ในไฟล์แม่แบบ")
            
            ws = wb[sheet_name]

            report_progress("  > กำลังกรอกวันที่ประมวลผล...")
            thai_date = self.date_formatter.format_thai_date(self.date.day(), self.date.month(), self.date.year())
            ws['D2'] = thai_date
            ws['D2'].font = self.date_formatter.date_font

            # 4. Prepare worksheet and fill data
            report_progress("  > กำลังเตรียมชีทงานและกรอกข้อมูล...")
            num_existing_files = len(processed_files_set)
            num_new_files = len(new_daily_files)
            
            self.worksheet_manager.prepare_worksheet(ws, num_new_files, num_existing_files)
            self.worksheet_manager.fill_data(ws, new_daily_files, new_data, self.inventory_code, num_existing_files)

            # 5. Add summary formulas
            report_progress("  > กำลังเพิ่มสูตรสรุป...")
            first_data_row = 5
            total_files = num_existing_files + num_new_files
            last_data_row = first_data_row + total_files - 1
            self.formula_manager.add_summary_formulas(ws, first_data_row, last_data_row)

            # 6. Handle cumulative sales
            report_progress("  > กำลังจัดการยอดขายสะสม...")
            self._handle_cumulative_sales(wb, ws, self.date.day(), last_data_row)
            
            report_progress("  > กำลังบันทึกสมุดงาน...")
            wb.save(self.template_path)

            # 7. Update state
            report_progress("  > กำลังอัปเดตสถานะการประมวลผล...")
            updated_processed_files = processed_files_set.union(set(new_daily_files))
            self.state_manager.save_processed_files(state_file_path, updated_processed_files)

            return Result.success(f"✅ ประมวลผลไฟล์ใหม่ {len(new_daily_files)} ไฟล์สำเร็จ")

        except PermissionError:
            return Result.error(f"❌ การเข้าถึงถูกปฏิเสธ ไฟล์แม่แบบอาจถูกเปิดอยู่:\n{self.template_path}\nกรุณาปิดไฟล์แล้วลองอีกครั้ง")
        except Exception as e:
            return Result.error(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิดระหว่างการประมวลผล Excel: {e}")

    def _handle_cumulative_sales(self, workbook, worksheet, day: int, last_data_row: int):
        """
        Handles the cumulative sales formula logic, carrying over from previous valid working days.
        """
        # Find the row containing "ยอดขายสะสม" (Cumulative Sales) in column D
        cumulative_sales_row = self.holiday_validator._find_row_by_value(worksheet, "D", "ยอดขายสะสม")
        if not cumulative_sales_row:
            return # No cumulative sales row found, nothing to do

        # Find the row containing "ยอดรวม" (Total) in column D
        total_row = self.holiday_validator._find_row_by_value(worksheet, "D", "ยอดรวม")
        current_day_total_value = worksheet[f"H{total_row}"].value if total_row else None

        if day == 1:
            # For the first day of the month, cumulative sales is just the current day's total
            if current_day_total_value is not None:
                worksheet[f"H{cumulative_sales_row}"] = current_day_total_value
        else:
            # For subsequent days, find the last valid working day's cumulative sales
            valid_prev_day_sheet, valid_prev_day_row = self.holiday_validator.find_last_valid_working_day(workbook, day)
            
            if valid_prev_day_sheet and valid_prev_day_row:
                # Found a valid previous working day, add its cumulative sales
                formula = f"=SUM(H5:H{last_data_row})+'{valid_prev_day_sheet}'!H{valid_prev_day_row}"
                worksheet[f"H{cumulative_sales_row}"] = formula
            else:
                # No valid previous day found, just sum current day's data
                formula = f"=SUM(H5:H{last_data_row})"
                worksheet[f"H{cumulative_sales_row}"] = formula
