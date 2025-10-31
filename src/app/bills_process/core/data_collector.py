import os
import re
from typing import List, Any, Tuple, Dict
from openpyxl import load_workbook
from .result import Result
from src.db.bill_config_queries import get_all_bill_configs

class DataCollector:
    """
    Handles collecting data from daily bill files based on configuration rules
    stored in the database.
    """
    
    def __init__(self):
        self.bill_configs = get_all_bill_configs()
    
    def collect_from_files(self, daily_files: List[str], target_dir: str, **kwargs) -> Result:
        """
        Collect data from all daily files based on the loaded bill configuration.
        
        Args:
            daily_files: List of daily bill file names.
            target_dir: Directory containing the daily files.
            
        Returns:
            Result: Contains a list of dictionaries (one per file) on success.
        """
        progress_callback = kwargs.get('progress_callback')
        def report_progress(msg):
            if progress_callback:
                progress_callback.emit(msg)

        if not self.bill_configs:
            return Result.error("❌ ไม่พบการตั้งค่าการประมวลผลบิล กรุณาตั้งค่าในหน้า 'การตั้งค่าการประมวลผลบิล'")

        try:
            collected_data = []
            total_files = len(daily_files)
            for i, fname in enumerate(daily_files):
                report_progress(f"  > กำลังรวบรวมข้อมูลจาก {fname} ({i+1}/{total_files})")
                file_path = os.path.join(target_dir, fname)
                file_result_dict = self._collect_from_single_file(file_path)
                collected_data.append(file_result_dict)
            
            return Result.success(collected_data)
            
        except Exception as e:
            return Result.error(f"❌ เกิดข้อผิดพลาดขณะรวบรวมข้อมูลจากไฟล์: {e}")
    
    def _collect_from_single_file(self, file_path: str) -> Dict[str, Any]:
        """
        Collect data from a single Excel file and return it as a dictionary
        where keys are the configured field names.
        """
        try:
            from src.utils.phone_number_extractor import extract_phone_numbers
            from src.utils.address_formatter import format_address_for_display

            wb_daily = load_workbook(file_path, data_only=True)
            ws_daily = wb_daily.active
            
            result_dict = {}
            
            for config in self.bill_configs:
                field_name = config.get('field_name')
                if not field_name:
                    continue
                
                value = self._extract_value_by_type(ws_daily, config)
                result_dict[field_name] = value
            
            # After collecting data, process the address field
            original_address = result_dict.get('receiver_address', '')
            if original_address:
                # 1. Extract phone number from the original address
                phone_numbers = extract_phone_numbers(original_address)
                result_dict['phone'] = phone_numbers

                # 2. Format the address and overwrite the original value
                formatted_address = format_address_for_display(original_address)
                result_dict['receiver_address'] = formatted_address

            return result_dict

        except Exception as e:
            print(f"Error processing file {os.path.basename(file_path)}: {e}")
            return {}

    def _extract_value_by_type(self, worksheet, config) -> Any:
        """
        Extract value from worksheet based on item type and configuration.
        """
        try:
            config_type = config['config_type'] # 0: By Cell, 1: By Column
            if config_type == 0: # By Cell
                return self._extract_by_cell(worksheet, config)
            elif config_type == 1: # By Column
                return self._extract_by_column(worksheet, config)
            else:
                return ""
        except Exception:
            return ""

    def _extract_by_cell(self, worksheet, config) -> Any:
        """Extract value using 'By Cell' method."""
        try:
            check_cell_address = config.get('check_text')
            expected_value = config.get('value')
            focus_cell_address = config.get('focus')

            if not all([check_cell_address, expected_value, focus_cell_address]):
                return ""

            actual_value = worksheet[check_cell_address].value
            
            # Compare values, stripping strings for safety
            if str(actual_value).strip() == str(expected_value).strip():
                return worksheet[focus_cell_address].value
            return ""
        except (KeyError, AttributeError):
            # Address might be invalid or cell doesn't exist
            return ""

    def _extract_by_column(self, worksheet, config) -> Any:
        """Extract value using 'By Column' method."""
        try:
            focus_col = config.get('focus', '').upper()
            check_val = config.get('check_text')
            collect_col = config.get('value', '').upper()
            
            if not all([focus_col, check_val, collect_col]):
                return ""

            for row in worksheet.iter_rows():
                focus_cell = None
                collect_cell = None
                for cell in row:
                    if cell.column_letter == focus_col:
                        focus_cell = cell
                    if cell.column_letter == collect_col:
                        collect_cell = cell
                
                if focus_cell and focus_cell.value is not None:
                    if str(focus_cell.value).strip() == str(check_val).strip():
                        return collect_cell.value if collect_cell else ""
            return ""
        except Exception:
            return ""
