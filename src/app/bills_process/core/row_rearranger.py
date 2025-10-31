import os
from openpyxl import load_workbook
from .result import Result
from .config_manager import ConfigManager

class RowRearranger:
    """
    Handles rearranging rows in the monthly template Excel file.
    """
    
    def __init__(self):
        self.config_manager = ConfigManager()

    def rearrange_template_rows(self, year: int, month: int, day: int) -> Result:
        """
        Rearranges rows in the template Excel file for the given date based on column L (platform)
        and then by column E (province/zone).
        """
        # 1. Get configuration
        config_result = self.config_manager.load_config()
        if not config_result.success:
            return config_result
        config = config_result.data
        base_path = config['base_path']

        # 2. Locate template file
        month_dir = os.path.join(base_path, f"Year_{year:04d}", f"Month_{month:02d}")
        template_file = os.path.join(month_dir, f"Monthly_Report_{month}_{year}.xlsx")
        
        if not os.path.exists(template_file):
            return Result.error(f"❌ ไม่พบไฟล์แม่แบบ:\n{template_file}")

        try:
            # 3. Open workbook and sheet
            wb = load_workbook(template_file)
            sheet_name = str(day)
            if sheet_name not in wb.sheetnames:
                return Result.error(f"❌ ไม่พบชีทชื่อ '{sheet_name}' ในไฟล์แม่แบบ")
            ws = wb[sheet_name]

            # 4. Read all data rows (starting from row 5, until first empty A cell or until formulas row)
            data_rows = []
            row_idx = 5
            while True:
                a_val = ws[f"A{row_idx}"].value
                # Stop if A is empty or if we hit a formula row (A cell is a formula)
                if a_val is None or (isinstance(a_val, str) and a_val.startswith("=")):
                    break
                
                # Read all columns A-N (1-14)
                row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, 15)]
                data_rows.append(row_data)
                row_idx += 1
            
            if not data_rows:
                return Result.info("ไม่มีแถวข้อมูลให้จัดเรียงใหม่")

            # 5. Sort by column L (index 11), then by column E (index 4)
            def sort_key(row_data_list):
                # Column L (index 11) for platform (Lazada/Shopee)
                l_val = (row_data_list[11] or '').strip().lower()
                if l_val == 'lazada':
                    l_order = 0
                elif l_val == 'shopee':
                    l_order = 1
                else:
                    l_order = 2
                
                # Column E (index 4) for province/zone
                e_val = str(row_data_list[4]) if row_data_list[4] is not None else ''
                return (l_order, e_val)
            
            data_rows.sort(key=sort_key)

            # 6. Write back sorted rows and update index in column A
            for idx, row_data in enumerate(data_rows):
                target_row_num = 5 + idx
                row_data[0] = idx + 1  # Update column A (index) to be sequential
                for col_num in range(1, 15): # Columns A-N
                    ws.cell(row=target_row_num, column=col_num).value = row_data[col_num-1]
            
            wb.save(template_file)
            return Result.success("✅ จัดเรียงแถวใหม่ตามแพลตฟอร์มและจังหวัด/โซนสำเร็จ")

        except PermissionError:
            return Result.error(f"❌ การเข้าถึงถูกปฏิเสธ ไฟล์แม่แบบอาจถูกเปิดอยู่:\n{template_file}\nกรุณาปิดไฟล์แล้วลองอีกครั้ง")
        except Exception as e:
            return Result.error(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิดระหว่างการจัดเรียงแถวใหม่: {e}")
