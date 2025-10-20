from openpyxl import load_workbook

def unmerge_cells_in_file(file_path: str):
    """
    Unmerges all merged cells in all worksheets of an Excel file.
    """
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        if sheet.merged_cells:
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_cell_range in merged_ranges:
                sheet.unmerge_cells(str(merged_cell_range))
    wb.save(file_path)
