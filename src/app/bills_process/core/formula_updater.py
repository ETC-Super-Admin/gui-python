import os
import re
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from .result import Result
from .config_manager import ConfigManager
from .excel_components.state_manager import StateManager # Needed to determine last_data_row
from .excel_components.holiday_validator import HolidayValidator # Needed for cumulative logic

class FormulaUpdater:
    """
    Handles updating specific formulas in the monthly template,
    especially for Lazada/Shopee totals and cumulative sales.
    """
    
    def __init__(self):
        self.config_manager = ConfigManager()
        self.state_manager = StateManager()
        self.holiday_validator = HolidayValidator()

    def update_formulas(self, year: int, month: int, day: int) -> Result:
        """
        Scans column E for 'Lazada' or 'Shopee' and updates related formulas.
        Also updates cumulative formulas for these platforms.
        """
        # 1. Get configuration
        config_result = self.config_manager.load_config()
        if not config_result.success:
            return config_result
        config = config_result.data
        base_path = config['base_path']

        # 2. Locate template file
        month_dir = os.path.join(base_path, f"Year_{year:04d}", f"Month_{month:02d}")
        template_file = os.path.join(month_dir, f"Monthly_Report_{month}_{year}.xlsx")
        
        if not os.path.exists(template_file):
            return Result.error(f"❌ ไม่พบไฟล์แม่แบบ:\n{template_file}")

        try:
            # 3. Open workbook and sheet
            wb = load_workbook(template_file)
            sheet_name = str(day)
            if sheet_name not in wb.sheetnames:
                return Result.error(f"❌ ไม่พบชีทชื่อ '{sheet_name}' ในไฟล์แม่แบบ")
            ws = wb[sheet_name]

            lazada_rows = []
            shopee_rows = []

            # Determine last_data_row from processed files state
            daily_bills_dir = os.path.join(
                base_path,
                f"Year_{year:04d}",
                f"Month_{month:02d}",
                "Daily_Bills",
                f"Day_{day}"
            )
            state_file_path = os.path.join(daily_bills_dir, ".processing_state.json")
            processed_files_set = self.state_manager.load_processed_files(state_file_path)
            num_processed_files = len(processed_files_set)

            first_data_row = 5
            if num_processed_files > 0:
                last_data_row = first_data_row + num_processed_files - 1
            else:
                last_data_row = first_data_row # Fallback if no data

            # Scan column E for Lazada/Shopee and update formulas
            for row_idx in range(1, ws.max_row + 1):
                e_cell_value = str(ws[f"E{row_idx}"].value or "").lower()
                
                if "lazada" in e_cell_value:
                    lazada_rows.append(row_idx)
                    self._update_platform_formulas(ws, row_idx, "Lazada", last_data_row, first_data_row)
                elif "shopee" in e_cell_value:
                    shopee_rows.append(row_idx)
                    self._update_platform_formulas(ws, row_idx, "Shopee", last_data_row, first_data_row)
            
            # Find 'Grand total' row
            row_grand_total = self.holiday_validator._find_row_by_value(ws, "E", "Grand total")

            # Update Grand Total and Cumulative formulas
            if shopee_rows and lazada_rows and row_grand_total:
                first_platform_row = min(shopee_rows[0], lazada_rows[0])
                last_platform_row = max(shopee_rows[-1], lazada_rows[-1])

                # Update Grand Total formulas (F, G, H)
                ws[f"F{row_grand_total}"] = f"=SUM(F{first_platform_row}:F{last_platform_row})"
                ws[f"G{row_grand_total}"] = f"=SUM(G{first_platform_row}:G{last_platform_row})"
                ws[f"H{row_grand_total}"] = f"=SUM(H{first_platform_row}:H{last_platform_row})"

                # Update Cumulative formulas (I) for each platform and Grand Total
                self._update_cumulative_platform_formulas(wb, ws, int(day), lazada_rows, shopee_rows, first_data_row)
                ws[f"I{row_grand_total}"] = f"=SUM(I{first_platform_row}:I{last_platform_row})"

            wb.save(template_file)
            return Result.success("✅ อัปเดตสูตรสำเร็จ")

        except PermissionError:
            return Result.error(f"❌ การเข้าถึงถูกปฏิเสธ ไฟล์แม่แบบอาจถูกเปิดอยู่:\n{template_file}\nกรุณาปิดไฟล์แล้วลองอีกครั้ง")
        except Exception as e:
            return Result.error(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิดระหว่างการอัปเดตสูตร: {e}")

    def _update_platform_formulas(self, ws, row_idx: int, platform: str, last_data_row: int, first_data_row: int):
        """Helper to update F, G, H formulas for a specific platform row."""
        ws[f"F{row_idx}"] = f"=COUNTIF(L{first_data_row}:L{last_data_row},\"{platform}\")"
        ws[f"G{row_idx}"] = f"=SUMIF(L{first_data_row}:L{last_data_row},\"{platform}\",G{first_data_row}:G{last_data_row})"
        ws[f"H{row_idx}"] = f"=SUMIF(L{first_data_row}:L{last_data_row},\"{platform}\",H{first_data_row}:H{last_data_row})"

    def _update_cumulative_platform_formulas(self, wb, ws, current_day: int, lazada_rows: List[int], shopee_rows: List[int], first_data_row: int):
        """
        Updates cumulative formulas (column I) for Lazada and Shopee rows.
        """
        for platform_row_idx in lazada_rows + shopee_rows:
            platform_name = str(ws[f"E{platform_row_idx}"].value or "").lower()
            
            if "lazada" in platform_name:
                keyword = "Lazada"
            elif "shopee" in platform_name:
                keyword = "Shopee"
            else:
                continue # Should not happen if rows are from lazada_rows/shopee_rows

            if current_day == 1:
                ws[f"I{platform_row_idx}"] = f"=H{platform_row_idx}"
            else:
                prev_day_sheet, prev_day_row = self._find_last_valid_platform_row(wb, current_day, keyword)
                if prev_day_sheet and prev_day_row:
                    ws[f"I{platform_row_idx}"] = f"=H{platform_row_idx}+'{prev_day_sheet}'!I{prev_day_row}"
                else:
                    ws[f"I{platform_row_idx}"] = f"=H{platform_row_idx}"

    def _find_last_valid_platform_row(self, workbook, current_day: int, keyword: str) -> Tuple[Optional[str], Optional[int]]:
        """
        Finds the last valid row for a specific platform (Lazada/Shopee) in previous days' sheets.
        """
        prev_day = current_day - 1
        while prev_day >= 1:
            prev_sheet_name = str(prev_day)
            if prev_sheet_name not in workbook.sheetnames:
                prev_day -= 1
                continue
            ws_prev = workbook[prev_sheet_name]
            
            # Find the row for the keyword in column E
            for row_idx in range(1, ws_prev.max_row + 1):
                e_cell_value = str(ws_prev[f"E{row_idx}"].value or "").lower()
                if keyword.lower() in e_cell_value:
                    # Check if column I of this row is valid (i.e., not empty or 0)
                    i_val = ws_prev[f"I{row_idx}"].value
                    if self.holiday_validator._is_valid_cumulative_value(i_val): # Re-use validator logic
                        return prev_sheet_name, row_idx
            prev_day -= 1
        return None, None
